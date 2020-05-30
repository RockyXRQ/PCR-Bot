from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Color, Font


class xlsHandle:
    wb = load_workbook(filename='PCR_Bot_Template.xlsx')  # 初始化读取xlsx文件


def wb_update():
    # 保存当前处理过的文件流 并且 更新全局文件流
    xlsHandle.wb.save('PCR_Bot_Template.xlsx')
    xlsHandle.wb = load_workbook(filename='PCR_Bot_Template.xlsx')


def xls_create_user(user_nickname: str):
    # 以用户昵称为sheet名新建sheet 并且 在该sheet首行填入用户昵称
    ws = xlsHandle.wb.copy_worksheet(xlsHandle.wb['Template'])
    ws.title = user_nickname
    ws['B1'] = user_nickname


def xls_damage_append(user_nickname: str, team_info: str, damage: int, day: int, boss: int):
    # 通过判断队伍构成是否相同判断补时刀
    def isSameTeam(orgTeam:str):
        if orgTeam:
            return set(team_info.split()) == set(str(orgTeam).split())
        else:
            return False


    ws_boss = xlsHandle.wb['Boss']
    ws_user = xlsHandle.wb[user_nickname]

    # 为去冗余定义信息填入函数
    def inner_damage_append(team_row: int, damage_row: int):
        isKill = False
        point = 0
        # 填入队伍信息
        ws_user.cell(team_row, 2+day).value = team_info

        # 如果不足以击杀Boss
        if damage <= ws_boss['B'+str(boss)].value:
            # 计算得分
            point += damage * ws_boss['C'+str(boss)].value
            # 更新Boss血量
            ws_boss['B'+str(boss)].value -= damage

        # 如果足以击杀Boss
        else:
            # 击杀状态为真
            isKill = True
            # 计算得分
            point = ws_boss['B'+str(boss)].value * \
                ws_boss['C'+str(boss)].value

            # 更新Boss血量
            ws_boss['B'+str(boss)].value = 0
        # 更新用户数据
        if ws_user.cell(damage_row, 2+day).value :
            ws_user.cell(damage_row, 2+day).value += point
        else:
            ws_user.cell(damage_row, 2+day).value = point
        # 若击杀则数据为红
        if isKill:
            ws_user.cell(damage_row, 2+day).font = Font(color=colors.RED)

        # 更新表格数据
        wb_update()
        return isKill

    # 获取对应用户的sheet
    ws = xlsHandle.wb[user_nickname]

    # 依次查看第一、二、三刀否存在数据 没有则填入
    if (not ws_user.cell(4, 2+day).value) or isSameTeam(ws_user.cell(4, 2+day).value):
        return inner_damage_append(4, 5)
    elif (not ws.cell(6, 2+day).value) or isSameTeam(ws.cell(6, 2+day).value):
        return inner_damage_append(6, 7)
    else:
        return inner_damage_append(8, 9)


def xls_on_tree(user_nickname: str):
    # 在对应用户的sheet上给 挂树次数 加一
    ws = xlsHandle.wb[user_nickname]
    ws.cell(5, 12).value += 1


def xls_save_tree(user_nickname: str):
    # 在对应用户的sheet上给 救树次 数加一
    ws = xlsHandle.wb[user_nickname]
    ws.cell(6, 12).value += 1


def xls_get_sheets():
    # 返回 sheet名称列表
    return xlsHandle.wb.sheetnames
