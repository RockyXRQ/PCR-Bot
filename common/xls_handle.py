from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Color, Font


class xlsHandle:
    wb = load_workbook(filename='PCR_Bot_Template.xlsx')  # 初始化读取xlsx文件


def wb_update():
    # 保存当前处理过的文件流 并且 更新全局文件流
    xlsHandle.wb.save('PCR_Bot_Template.xlsx')
    xlsHandle.wb = load_workbook(filename='PCR_Bot_Template.xlsx')


def xls_create_user(user_nickname: str):
    # 以用户昵称为sheet名新建sheet 并且 在该sheet首行填入用户昵称
    ws = xlsHandle.wb.copy_worksheet(xlsHandle.wb['Template'])
    ws.title = user_nickname
    ws['B1'] = user_nickname


def xls_damage_append(user_nickname: str, team_info: str, damage: int, day: int, isKill: bool):
    ft = Font(color=colors.RED)

    # 为去冗余定义信息填入函数
    def inner_damage_append(team_row: int, damage_row: int):
        ws.cell(team_row, 2+day).value = team_info
        ws.cell(damage_row, 2+day).value = damage
        if isKill:
            ws.cell(damage_row, 2+day).font = ft

    # 获取对应用户的sheet
    ws = xlsHandle.wb[user_nickname]

    # 依次查看第一、二、三刀否存在数据 没有则填入
    if not ws.cell(4, 2+day).value:
        inner_damage_append(4, 5)
    elif not ws.cell(6, 2+day).value:
        inner_damage_append(6, 7)
    else:
        inner_damage_append(8, 9)

    # 更新表格数据
    wb_update()


def xls_on_tree(user_nickname: str):
    # 在对应用户的sheet上给 挂树次数 加一
    ws = xlsHandle.wb[user_nickname]
    ws.cell(5, 12).value += 1


def xls_save_tree(user_nickname: str):
    # 在对应用户的sheet上给 救树次 数加一
    ws = xlsHandle.wb[user_nickname]
    ws.cell(6, 12).value += 1


def xls_get_sheets():
    # 返回 sheet名称列表
    return xlsHandle.wb.sheetnames


def xls_get_boss_list():
    # 返回Boss名称与血量组成的数组
    ws = xlsHandle.wb['Boss']
    return [
        [ws['A1'], ws['B1'].value],
        [ws['A2'], ws['B2'].value],
        [ws['A3'], ws['B3'].value],
        [ws['A4'], ws['B4'].value],
        [ws['A5'], ws['B5'].value]
    ]


def xls_update_boss_list(boss: int, damage: int):
    # 对Boss sheet开始处理
    ws = xlsHandle.wb['Boss']
    HP = ws['B'+str(boss)].value

    # 补时刀血量处理
    if damage <= ws['B'+str(boss)].value:
        ws['B'+str(boss)].value = max(0, HP - damage)
    else:
        remain = damage - HP
        ws['B'+str(boss)].value = 0
        ws['B'+str((boss % 5) + 1)].value -= remain


def xls_restore_boss_list():
    # 每周回打完后刷新全部Boss血量数据
    ws = xlsHandle.wb['Boss']
    ws['B1'] = 6000000
    ws['B2'] = 8000000
    ws['B3'] = 10000000
    ws['B4'] = 12000000
    ws['B5'] = 20000000
